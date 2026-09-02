"""
Script para truncar la BD y cargar datos 2026 desde el Excel.
Ejecutar desde backend/:  python load_excel.py
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from datetime import datetime
import openpyxl
from sqlalchemy import text
from app.database import engine, SessionLocal

EXCEL = os.path.join(
    os.path.dirname(__file__), "..",
    "ingresos fix fast (version 1) (version 1).xlsb (1).xlsx"
)

# ── Columnas de egresos en las hojas mensuales (nombre_header → categoría) ──
EGRESO_COLS = {
    "Sueldo Lisette":      "Sueldo Lisette",
    "sueldo gilberto":     "Sueldo Gilberto",
    "chucho":              "Sueldo Chucho",
    "Alejandro":           "Sueldo Alejandro",
    "Miller":              "Sueldo Miller",
    "Arreglos otros":      "Arreglos terceros",
    "Prestamo senor Gilberto": "Préstamo Sr. Gilberto",
    "onces y almuerzo":    "Alimentación",
    "GASTOS LOCAL":        "Gastos del local",
    "GASTOS PERSONALES":   "Gastos personales",
}

# Columnas de ingreso en las hojas mensuales → metodo_pago en movimientos_caja
INGRESO_COLS = {
    "producido efectivo":  "efectivo",
    "Total producido":     "efectivo",
    "Datafono Bruto":      "tarjeta",
    "Nequi Neto":          "nequi",
    "Bancolombia":         "transferencia",
}


def safe_float(v):
    """Convierte a float tolerando espacios, strings basura y None."""
    try:
        return float(v) if v and str(v).strip() else 0
    except (ValueError, TypeError):
        return 0


def truncate_all(session):
    """Trunca todas las tablas excepto usuarios (ligada a auth.users)."""
    tables = [
        "used_parts",
        "movimientos_inventario",
        "pagos",
        "historial_estados",
        "movimientos_caja",
        "ordenes",
        "clientes",
        "productos",
        "categorias_inventario",
        "proveedores",
        "categorias",
    ]
    for t in tables:
        session.execute(text(f'TRUNCATE TABLE "{t}" RESTART IDENTITY CASCADE'))
    session.commit()
    print(f"✓ Truncadas {len(tables)} tablas")


def clean_phone(telefono):
    """Limpia teléfono: quita .0 de floats, deja strings como están."""
    if not telefono:
        return None
    t = str(telefono).strip()
    if t.endswith(".0"):
        t = t[:-2]
    return t or None


def load_ordenes(wb, session):
    """Carga clientes y órdenes desde 'ingresos a taller' (solo 2026).
    NO crea pagos — los ingresos se cargan desde las hojas mensuales."""
    ws = wb["ingresos a taller"]
    rows = list(ws.iter_rows(values_only=True))

    # ── 1. Crear clientes únicos ──
    client_map = {}  # (nombre, telefono) → client_id
    for row in rows[1:]:
        fecha, recibo, nombre, telefono = row[0], row[1], row[2], row[3]
        if not fecha or not hasattr(fecha, "year") or fecha.year != 2026:
            continue
        if nombre is None:
            continue
        nombre = str(nombre).strip()
        telefono = clean_phone(telefono)
        key = (nombre, telefono or "")
        if key not in client_map:
            client_map[key] = None

    client_rows = []
    for nombre, telefono in client_map:
        client_rows.append({"nombre": nombre, "telefono": telefono or None})

    for row_data in client_rows:
        result = session.execute(
            text(
                'INSERT INTO clientes (nombre, telefono) '
                'VALUES (:nombre, :telefono) RETURNING id'
            ),
            row_data,
        )
        cid = result.scalar()
        key = (row_data["nombre"], row_data["telefono"] or "")
        client_map[key] = cid
    session.commit()
    print(f"✓ {len(client_rows)} clientes creados")

    # ── 2. Crear órdenes ──
    orden_count = 0
    for row in rows[1:]:
        fecha, recibo, nombre, telefono, tipo_arreglo = row[0], row[1], row[2], row[3], row[4]
        valor_excel, abono_excel, saldo_excel, fecha_entrega = row[5], row[6], row[7], row[8]

        if not fecha or not hasattr(fecha, "year") or fecha.year != 2026:
            continue
        if nombre is None:
            continue

        nombre = str(nombre).strip()
        telefono = clean_phone(telefono)
        key = (nombre, telefono or "")
        cliente_id = client_map.get(key)
        if not cliente_id:
            continue

        valor_num = safe_float(valor_excel)
        abono_num = safe_float(abono_excel)
        saldo_num = safe_float(saldo_excel)

        # ── Corregir valor cuando "Valor del arreglo" está vacío ──
        # En el Excel, si no llenaron la columna Valor pero sí Saldo,
        # el saldo ES el precio total (nada pagado aún).
        # Si hay abono, el precio total = abono + saldo.
        if valor_num == 0:
            valor_num = abono_num + saldo_num

        # Recalcular saldo si falta en el Excel
        if saldo_num == 0 and valor_num > 0 and abono_num > 0:
            saldo_num = max(0, valor_num - abono_num)

        # Determinar estado (con mayúscula — formato canónico del app)
        if fecha_entrega and hasattr(fecha_entrega, "year"):
            estado = "Entregado"
        elif valor_num > 0 and saldo_num == 0 and abono_num > 0:
            estado = "Entregado"
        else:
            estado = "Pendiente"

        fe = None
        if fecha_entrega and hasattr(fecha_entrega, "year"):
            fe = fecha_entrega

        numero_orden = str(int(recibo)) if recibo else None

        session.execute(
            text(
                'INSERT INTO ordenes '
                '(cliente_id, numero_orden, problema, valor, saldo, estado, '
                ' fecha_ingreso, fecha_entrega, created_at) '
                'VALUES (:cliente_id, :numero_orden, :problema, :valor, :saldo, '
                ' :estado, :fecha_ingreso, :fecha_entrega, :fecha_ingreso)'
            ),
            {
                "cliente_id": cliente_id,
                "numero_orden": numero_orden,
                "problema": str(tipo_arreglo).strip() if tipo_arreglo else None,
                "valor": valor_num,
                "saldo": saldo_num,
                "estado": estado,
                "fecha_ingreso": fecha,
                "fecha_entrega": fe,
            },
        )
        orden_count += 1

    session.commit()
    print(f"✓ {orden_count} órdenes creadas")


def load_movimientos(wb, session):
    """Carga ingresos y egresos de las hojas mensuales 2026 como movimientos_caja."""
    sheets_2026 = [
        "ENERO 2026", "FEBRERO 2026", "MARZO 2026", "ABRIL 2026",
        "MAYO 2026", "JUNIO 2026", "JULIO 2026", "AGOSTO 2026",
    ]

    # Crear categorías
    cat_nombres = set(EGRESO_COLS.values()) | {"Ingresos taller"}
    for cat_nombre in cat_nombres:
        tipo = "ingreso" if cat_nombre == "Ingresos taller" else "egreso"
        session.execute(
            text(
                "INSERT INTO categorias (nombre, tipo) VALUES (:nombre, :tipo) "
                "ON CONFLICT DO NOTHING"
            ),
            {"nombre": cat_nombre, "tipo": tipo},
        )
    session.commit()

    ingreso_count = 0
    egreso_count = 0

    for sheet_name in sheets_2026:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]

        # Mapear columnas
        egreso_map = {}   # idx → categoria
        ingreso_map = {}  # idx → metodo_pago
        for idx, col_name in enumerate(header):
            if not col_name:
                continue
            name = str(col_name).strip()
            if name in EGRESO_COLS:
                egreso_map[idx] = EGRESO_COLS[name]
            if name in INGRESO_COLS:
                ingreso_map[idx] = INGRESO_COLS[name]

        for row in rows[1:]:
            fecha = row[0]
            if not fecha or not hasattr(fecha, "year"):
                continue

            # ── Ingresos ──
            for idx, metodo in ingreso_map.items():
                if idx < len(row) and row[idx] is not None:
                    valor = safe_float(row[idx])
                    if valor <= 0:
                        continue
                    session.execute(
                        text(
                            'INSERT INTO movimientos_caja '
                            '(tipo, categoria, valor, metodo_pago, descripcion, created_at) '
                            'VALUES (:tipo, :cat, :valor, :metodo, :desc, :fecha)'
                        ),
                        {
                            "tipo": "ingreso",
                            "cat": "Ingresos taller",
                            "valor": valor,
                            "metodo": metodo,
                            "desc": f"Ingreso {metodo} — {sheet_name}",
                            "fecha": fecha,
                        },
                    )
                    ingreso_count += 1

            # ── Egresos ──
            for idx, categoria in egreso_map.items():
                if idx < len(row) and row[idx] is not None:
                    valor = safe_float(row[idx])
                    if valor <= 0:
                        continue
                    session.execute(
                        text(
                            'INSERT INTO movimientos_caja '
                            '(tipo, categoria, valor, metodo_pago, descripcion, created_at) '
                            'VALUES (:tipo, :cat, :valor, :metodo, :desc, :fecha)'
                        ),
                        {
                            "tipo": "egreso",
                            "cat": categoria,
                            "valor": valor,
                            "metodo": "efectivo",
                            "desc": f"{categoria} — {sheet_name}",
                            "fecha": fecha,
                        },
                    )
                    egreso_count += 1

        session.commit()

    print(f"✓ {ingreso_count} movimientos de ingreso creados")
    print(f"✓ {egreso_count} movimientos de egreso creados")


def main():
    print("Cargando Excel…")
    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    session = SessionLocal()
    try:
        print("\n── Truncando tablas ──")
        truncate_all(session)

        print("\n── Cargando órdenes 2026 ──")
        load_ordenes(wb, session)

        print("\n── Cargando movimientos 2026 (ingresos + egresos) ──")
        load_movimientos(wb, session)

        # ── Resumen final ──
        print("\n── Verificación ──")
        for tabla in ["clientes", "ordenes", "pagos", "movimientos_caja"]:
            n = session.execute(text(f"SELECT count(*) FROM {tabla}")).scalar()
            print(f"  {tabla}: {n}")
        ing = session.execute(text(
            "SELECT COALESCE(SUM(valor),0) FROM movimientos_caja WHERE tipo='ingreso'"
        )).scalar()
        egr = session.execute(text(
            "SELECT COALESCE(SUM(valor),0) FROM movimientos_caja WHERE tipo='egreso'"
        )).scalar()
        val = session.execute(text(
            "SELECT COALESCE(SUM(valor),0) FROM ordenes"
        )).scalar()
        print(f"  Total ingresos (mov_caja): ${ing:,.0f}")
        print(f"  Total egresos (mov_caja):  ${egr:,.0f}")
        print(f"  Total valor órdenes:       ${val:,.0f}")

        print("\n✅ Carga completa")
    except Exception as e:
        session.rollback()
        print(f"\n❌ Error: {e}")
        raise
    finally:
        session.close()


if __name__ == "__main__":
    main()
